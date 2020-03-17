#required library
import openpyxl as xl  
import matplotlib.pyplot as plt



#calculate percentage for your excel sheet
def perc_calc(str1,n,per_subjectmarks,str2):
    wb=xl.load_workbook(str1) #load your workbook
    sheet=wb['Sheet1']
    perc = sheet.cell(1, n+2)
    perc.value = 'perc %' # name of new colume which is added to new excel sheet 'perc %'
    tot=n*per_subjectmarks #  total number of marks 
    for row in range(2,sheet.max_row+1):
        sum1=0
        tot_marks=sheet.cell(row,n+2)

        for col in range(2,n+2):
            val=sheet.cell(row,col)
            sum1 +=val.value
            tot_marks.value=sum1
        per=(sum1/tot)*100
        perc=sheet.cell(row,n+2)
        perc.value=per
        wb.save(str2) # save your file to the provided loction
#give remarks to your student
def cong(n):
    wb=xl.load_workbook(str2) #load your workbook
    sheet=wb['Sheet1']
    for row in range(2,sheet.max_row+1):
        out=sheet.cell(row,n+2)
        san=sheet.cell(row,n+3)
        if((out.value)<=50):
            san.value='NEED TO IMPROVE'
        elif(out.value>50 and out.value<74.9):
            san.value='great score you can do better'
        elif(out.value>=75 and out.value<85):
            san.value='you r charme'
        else:
            san.value='Go and play health is also imortant'
    wb.save(str2)


#searching for student percentage
def serch():
    n=int(input('enter the student roll no '))
    select=sheet.cell(n+1,n+2)
    print('student marks' ,select.value)


#add pie chart to your data
def add_pie(str2,str3):
    wb2 = xl.load_workbook(str2)
    sheet2 = wb2['Sheet1']
    cou50 = 0
    cou75 = 0
    cou85 = 0
    cou86 = 0
    labes = ['below 50 ', 'between 50 to 75', 'between 75 to 85', 'above 85'] #labels of pie chart
    for row in range(2, sheet2.max_row + 1):
        cel = sheet2.cell(row, n+2)
        if (cel.value <= 50):
            cou50 += 1
        elif (cel.value > 50 and cel.value <= 75):
            cou75 += 1
        elif (cel.value > 75 and cel.value <= 85):
            cou85 += 1
        else:
            cou86 += 1
    values = [cou50, cou75, cou85, cou86]

    plt.figure()
    plt.title('student chart')
    explode = (0, 0, 0, 0.1)
    plt.pie(values, labels=labes, autopct='%1.1f%%', shadow=True, startangle=90, explode=explode)
    
    str3 =str3+'\performance.png' # name of pie chart file in png formate 
    plt.savefig(str3)
    plt.show
    

print('YOUR EXCEL SHEET SHOULD BE IN THE FORMATE')
print('roll no |subj1 |subject2 |subject3 |.....subjectn')
str1=input("Provide the location of excel file with xlxs extension for exaple ")
str2=input("Provide the location where u have to stroe the file excel file with xlxs extension for exaple ")
str3=input("provide where u want to store pie chart")
n=int(input('provide the no of subject '))
per_subjectmarks=int(input("provide per subject marks"))

perc_calc(str1,n,per_subjectmarks,str2)
cong(n)
add_pie(str2,str3)

print('your result will be store in excel sheet & pie graph is created ') 


    

